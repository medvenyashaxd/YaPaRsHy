import os
import xlsxwriter
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from selenium.common import TimeoutException
from selenium.webdriver.chrome.options import Options
from Locators.locators import LocatorsYandexSearchCatalog
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.support import expected_conditions as Ec


class Parser:

    def executor(self, search_data, locators=LocatorsYandexSearchCatalog):

        driver_path = ('webdriver\chromedriver.exe')
        options = Options()
        options.add_argument(r"user-data-dir=C:\Users\xmedv\AppData\Local\Google\Chrome\User Data\Default")

        driver = webdriver.Chrome(executable_path=driver_path, options=options)

        main_url_search = 'https://partner.market.yandex.ru/business/1030699/assortment?campaignId=22031686&activeTab=mappings&mappingChangePeriod=all&mappingChecking=processed'

        driver.get(main_url_search)

        len_search_data = len(search_data)
        for l in range(len_search_data):
            Wait(driver, timeout=20).until(Ec.presence_of_element_located(locator=locators.SEARCH_FIELD)).send_keys(
                search_data[l])

            try:
                Wait(driver, timeout=5, poll_frequency=0.1).until(Ec.text_to_be_present_in_element(locator=locators.SKU, text_=search_data[l]))

                with open('site.html', 'w', encoding='utf-8') as file:
                    file.write(driver.page_source)
                file.close()

                Wait(driver, timeout=20).until(Ec.visibility_of_element_located(locator=locators.SEARCH_FIELD)) \
                    .send_keys(u'\ue009' + u'\ue003')

                self.parser(search_data[l])

            except TimeoutException:
                with open('site.html', 'w', encoding='utf-8') as file:
                    file.write(driver.page_source)
                file.close()

                Wait(driver, timeout=20).until(Ec.visibility_of_element_located(locator=locators.SEARCH_FIELD)) \
                    .send_keys(u'\ue009' + u'\ue003')

                self.parser(search_data[l])

    def parser(self, search_data):

        with open('site.html', 'r', encoding='utf-8') as file:
            html = file.read()
        file.close()

        bs = BeautifulSoup(html, 'lxml')

        check_nothing_found = bs.find('div', class_='___root_k72io_1 __use--inline_k72io_1')

        if check_nothing_found is None:
            our_products_found = bs.find_all('div', class_='style-root___oDOxj')
            bindings = bs.find_all('span', class_='style-checkingStatus___FC806')
            cards_market = bs.find_all('span', class_='style-linkWrapper___H264c')

            skus_list = []
            skus = bs.find_all('span', class_='style-sku___vuyKz')
            for sku in skus:
                skus_list.append(sku.text)

            for our_product in our_products_found:
                sku = our_product.find('span', 'style-sku___vuyKz').text
                if sku == search_data:
                    index = skus_list.index(sku)
                    product = our_product.find('a', '___Clickable___fcJVD').text
                    card_market_name = cards_market[index].find('span',
                                                           class_='style-linkContnent___IRaY_ style-linkText___RWnjN').text
                    href = cards_market[index].find('a', class_='___Clickable___fcJVD').get('href').replace('//', '')
                    binding = bindings[index].find('span').text

                    return self.save_data_pars(sku, product, card_market_name, href, binding)

        else:
            return self.save_data_pars(search_data, '???? ???????????? ?????????????? ???????????? ???? ??????????????',
                                       '???? ???????????? ?????????????? ???????????? ???? ??????????????',
                                       '???? ???????????? ?????????????? ???????????? ???? ??????????????',
                                       '???? ???????????? ?????????????? ???????????? ???? ??????????????',)

    def save_data_pars(self, sku, product, card_market_name, href, binding):
        file_name = 'privyazki/privyazki.xlsx'

        if os.path.exists(file_name):
            xl_file = load_workbook(file_name)
            page = xl_file['privyazki']
            page.append([sku, product, card_market_name, href, binding])

            xl_file.save(file_name)
            xl_file.close()

        else:
            book = xlsxwriter.Workbook(file_name)  # ?????????????? ???????? Exel
            page = book.add_worksheet('privyazki')

            row = 0

            page.set_column('A:A', 12)
            page.set_column('B:B', 45)
            page.set_column('C:C', 75)
            page.set_column('D:D', 60)
            page.set_column('E:E', 15)

            page.write(0, 0, 'SKU')
            page.write(0, 0 + 1, '?????? ??????????')
            page.write(0, 0 + 2, '???????????????? ??????????????:')
            page.write(0, 0 + 3, '???????????? ???? ????????????????')
            page.write(0, 0 + 4, '????????????????')
            row += 1

            book.close()

            xl_file = load_workbook(file_name)
            page = xl_file['privyazki']

            page.append([sku, product, card_market_name, href, binding])

            xl_file.save(file_name)
            xl_file.close()


if __name__ == '__main__':
    data = (
        '30501DBS',
        '43420DBS',
    )

    Parser().executor(data)
